"""
import_service.py — Bulk import service for CSV/XLSX files.
Supports Client, Product, Material resources with validation and side effects.
"""
import csv
import io
import os
import re
from datetime import datetime
from typing import List, Dict, Tuple

from fastapi import HTTPException, UploadFile
from sqlalchemy.orm import Session
from pydantic import ValidationError
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException

import models
import schemas


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
MAX_ROWS_PARSE = 10000
MAX_ROWS_PREVIEW = 5000

# Spanish/normalized header -> field name mapping
HEADER_MAP = {
    "clients": {
        "nombre": "name",
        "email": "email",
        "telefono": "phone",
        "direccion": "address",
        "cuit": "tax_id",
        "condicion_iva_receptor_id": "condicion_iva_receptor_id",
    },
    "products": {
        "sku": "sku",
        "nombre": "name",
        "precio": "price",
        "stock": "stock",
        "stock_minimo": "stock_minimo",
    },
    "materials": {
        "sku": "sku",
        "nombre": "name",
        "categoria": "category",
        "stock_actual": "current_stock",
        "costo_unitario": "unit_cost",
        "precio_unitario": "unit_price",
        "stock_minimo": "stock_minimo",
    },
}

# Accepted English field names (mapped to themselves)
VALID_FIELD_NAMES = {
    "clients": {"name", "email", "phone", "address", "tax_id", "condicion_iva_receptor_id"},
    "products": {"sku", "name", "price", "stock", "stock_minimo"},
    "materials": {"sku", "name", "category", "current_stock", "unit_cost", "unit_price", "stock_minimo"},
}

SCHEMA_MAP = {
    "clients": schemas.ClientCreate,
    "products": schemas.ProductCreate,
    "materials": schemas.MaterialCreate,
}

TEMPLATE_HEADERS = {
    "clients": ["nombre", "email", "telefono", "direccion", "cuit", "condicion_iva_receptor_id"],
    "products": ["sku", "nombre", "precio", "stock", "stock_minimo"],
    "materials": ["sku", "nombre", "categoria", "stock_actual", "costo_unitario", "precio_unitario", "stock_minimo"],
}

VALID_RESOURCES = {"clients", "products", "materials"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalize_header(header: str) -> str:
    h = header.strip().lower()
    h = h.replace(" ", "_").replace("-", "_")
    h = "".join(c for c in h if c.isalnum() or c == "_")
    return h


def _normalize_number(value: str) -> str:
    """Convert locale-formatted numbers (e.g. '4,5' -> '4.5')."""
    if isinstance(value, str):
        value = value.strip()
        if re.match(r'^-?\d+,\d+$', value):
            value = value.replace(",", ".")
    return value


def _detect_delimiter(sample: str) -> str:
    lines = sample.split("\n")
    if not lines:
        return ","
    first = lines[0]
    counts = {
        ",": first.count(","),
        ";": first.count(";"),
        "\t": first.count("\t"),
    }
    best = max(counts, key=counts.get)
    return best if counts[best] > 0 else ","


def _empty_to_none(value):
    """Convert empty/whitespace-only string to None."""
    if isinstance(value, str) and value.strip() == "":
        return None
    return value


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def _parse_csv(content: bytes, warnings: List[str]) -> List[Dict]:
    """Parse CSV content -> list of dicts with original header keys."""
    raw = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            raw = content.decode(encoding)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if raw is None:
        raise HTTPException(status_code=400, detail="No se pudo detectar la codificacion del archivo CSV")

    delimiter = _detect_delimiter(raw)
    reader = csv.DictReader(io.StringIO(raw), delimiter=delimiter)

    if not reader.fieldnames:
        raise HTTPException(
            status_code=400,
            detail="No se detectaron headers. Verifique que la primera fila contenga los nombres de columna",
        )

    rows = []
    for row in reader:
        cleaned = {}
        for k, v in row.items():
            if k is None:
                continue
            cleaned[k] = _normalize_number(v) if isinstance(v, str) else v
        rows.append(cleaned)

    return rows


def _parse_xlsx(content: bytes, warnings: List[str]) -> List[Dict]:
    """Parse XLSX content (first sheet only) -> list of dicts with original header keys."""
    wb = None
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    except InvalidFileException:
        raise HTTPException(status_code=400, detail="El archivo Excel no es valido o esta danado")
    except Exception as e:
        msg = str(e).lower()
        if "password" in msg or "protected" in msg:
            raise HTTPException(
                status_code=400,
                detail="El archivo esta protegido con contrasena. Quite la proteccion e intente nuevamente",
            )
        raise HTTPException(status_code=400, detail=f"Error al leer archivo Excel: {str(e)}")

    sheet = wb.active
    if sheet is None:
        wb.close()
        raise HTTPException(status_code=400, detail="El archivo Excel no tiene hojas")

    header_row = next(sheet.iter_rows(values_only=True), None)
    if not header_row:
        wb.close()
        raise HTTPException(
            status_code=400,
            detail="No se detectaron headers. Verifique que la primera fila contenga los nombres de columna",
        )

    headers = []
    for h in header_row:
        h_str = str(h).strip() if h is not None else ""
        if h_str:
            headers.append(h_str)

    if not headers:
        wb.close()
        raise HTTPException(
            status_code=400,
            detail="No se detectaron headers. Verifique que la primera fila contenga los nombres de columna",
        )

    rows = []
    for row in sheet.iter_rows(values_only=True):
        if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
            continue

        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                val = row[i]
                if isinstance(val, str):
                    val = val.strip()
                elif val is None:
                    val = ""
                row_dict[header] = val
        rows.append(row_dict)

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Column mapping
# ---------------------------------------------------------------------------

def _build_column_mapping(headers: List[str], resource: str) -> Tuple[Dict[str, str], List[str]]:
    """Build (mapped: {original_header: field_name}, ignored: [original_headers])."""
    header_map = HEADER_MAP[resource]
    valid_fields = VALID_FIELD_NAMES[resource]

    mapped = {}
    ignored = []

    for header in headers:
        normalized = _normalize_header(header)
        if not normalized:
            continue
        if normalized in header_map:
            mapped[header] = header_map[normalized]
        elif normalized in valid_fields:
            mapped[header] = normalized
        else:
            ignored.append(header)

    return mapped, ignored


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_file(file: UploadFile) -> Tuple[List[Dict], List[str]]:
    """Parse uploaded file (CSV or XLSX).

    Returns (rows, warnings).
    Each row is a dict with original (un-normalized) header keys.
    """
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()

    if ext not in (".csv", ".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="Formato no soportado. Use CSV o Excel (.xlsx)",
        )

    content = file.file.read()

    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail="El archivo excede el tamano maximo de 10MB",
        )

    if len(content) == 0:
        raise HTTPException(status_code=400, detail="El archivo esta vacio")

    warnings = []

    if ext == ".csv":
        data_rows = _parse_csv(content, warnings)
    else:
        data_rows = _parse_xlsx(content, warnings)

    # Filter completely empty rows
    data_rows = [
        row
        for row in data_rows
        if any(
            isinstance(v, str) and v.strip() != "" or not isinstance(v, str)
            for v in row.values()
        )
    ]

    if not data_rows:
        raise HTTPException(status_code=400, detail="El archivo no contiene filas de datos")

    if len(data_rows) > MAX_ROWS_PARSE:
        warnings.append(
            f"El archivo tiene {len(data_rows)} filas. Procesando primeras {MAX_ROWS_PARSE}."
        )
        data_rows = data_rows[:MAX_ROWS_PARSE]

    return data_rows, warnings


# ---------------------------------------------------------------------------
# Error formatting
# ---------------------------------------------------------------------------

FIELD_LABELS = {
    "name": "nombre",
    "email": "email",
    "phone": "telefono",
    "address": "direccion",
    "tax_id": "CUIT",
    "condicion_iva_receptor_id": "condicion IVA",
    "sku": "SKU",
    "price": "precio",
    "stock": "stock",
    "stock_minimo": "stock minimo",
    "category": "categoria",
    "current_stock": "stock actual",
    "unit_cost": "costo unitario",
    "unit_price": "precio unitario",
}


def field_label(field: str) -> str:
    """Map a field name to its Spanish label."""
    return FIELD_LABELS.get(field, field)


def format_pydantic_error(error: Dict, value) -> str:
    """Format a Pydantic ValidationError item into a Spanish error message."""
    loc = error.get("loc", [])
    field = str(loc[0]) if loc else "campo"
    etype = error.get("type", "")

    if etype == "missing":
        return f"El campo '{field_label(field)}' es obligatorio"
    if etype in ("string_too_short",):
        return f"El campo '{field_label(field)}' es obligatorio"
    if etype in ("string_too_long",):
        return f"El campo '{field_label(field)}' es demasiado largo"
    if etype == "value_error":
        return f"El valor '{value}' no es valido para '{field_label(field)}'"
    if "float_parsing" in etype or ("float" in etype and "type" in etype):
        return f"El campo '{field_label(field)}' debe ser un numero"
    if "int_parsing" in etype or "int_from_float" in etype:
        return f"El campo '{field_label(field)}' debe ser un numero entero"
    if etype in ("greater_than_equal", "less_than_equal"):
        ctx_val = error.get("ctx", {}) or {}
        if "ge" in ctx_val:
            return f"El campo '{field_label(field)}' debe ser mayor o igual a {ctx_val['ge']}"
        if "le" in ctx_val:
            return f"El campo '{field_label(field)}' debe ser menor o igual a {ctx_val['le']}"
        return f"El campo '{field_label(field)}' no cumple con la validacion numerica"
    if etype == "string_pattern_mismatch":
        if field == "tax_id":
            return f"El CUIT '{value}' no tiene un formato valido (use XX-XXXXXXXX-X)"
        return f"El campo '{field_label(field)}' no tiene un formato valido"
    if etype == "email":
        return f"El email '{value}' no tiene un formato valido"

    # Fallback
    return f"Error en campo '{field_label(field)}': {error.get('msg', 'valor invalido')}"


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_rows(rows: List[Dict], resource: str, db: Session) -> Dict:
    """Validate parsed rows against Pydantic schema.

    Returns::

        {
            "valid_rows": [{"row": idx, "data": {field: value}}],
            "error_rows": [{"row": idx, "data": {field: value}, "errors": {field: msg}}],
            "columns_mapped": {original_header: field_name},
            "columns_ignored": [original_header],
        }
    """
    schema_class = SCHEMA_MAP[resource]

    # Build column mapping from first row
    headers = list(rows[0].keys()) if rows else []
    column_map, ignored_columns = _build_column_mapping(headers, resource)

    valid_rows = []
    error_rows = []
    seen_skus: set = set()

    # Gather existing SKUs from DB for products / materials
    existing_skus: set = set()
    if resource == "products":
        existing = db.query(models.Product.sku).filter(models.Product.sku.isnot(None)).all()
        existing_skus = {s[0] for s in existing if s[0]}
    elif resource == "materials":
        existing = db.query(models.Material.sku).filter(models.Material.sku.isnot(None)).all()
        existing_skus = {s[0] for s in existing if s[0]}

    for row_idx, row in enumerate(rows):
        # Map column names to field names
        mapped_data: Dict = {}
        for key, value in row.items():
            if key in column_map:
                mapped_data[column_map[key]] = _empty_to_none(value)
            elif key in VALID_FIELD_NAMES[resource]:
                mapped_data[key] = _empty_to_none(value)

        field_errors: Dict[str, str] = {}

        # Validate with Pydantic schema
        try:
            validated = schema_class(**{k: v for k, v in mapped_data.items() if v is not None})
            clean_data = validated.model_dump()
        except ValidationError as e:
            for err in e.errors():
                f = str(err["loc"][0]) if err.get("loc") else "unknown"
                field_errors[f] = format_pydantic_error(err, mapped_data.get(f, ""))

        # SKU uniqueness check
        if not field_errors and resource in ("products", "materials"):
            sku = mapped_data.get("sku")
            if sku is not None:
                sku_str = str(sku).strip()
                if sku_str in seen_skus:
                    field_errors["sku"] = f"El SKU '{sku_str}' esta duplicado en el archivo"
                elif sku_str in existing_skus:
                    label = "producto" if resource == "products" else "material"
                    field_errors["sku"] = f"El SKU '{sku_str}' ya existe en otro {label}"
                else:
                    seen_skus.add(sku_str)

        if field_errors:
            error_rows.append({"row": row_idx, "data": mapped_data, "errors": field_errors})
        else:
            valid_rows.append({"row": row_idx, "data": clean_data})

    return {
        "valid_rows": valid_rows,
        "error_rows": error_rows,
        "columns_mapped": column_map,
        "columns_ignored": ignored_columns,
    }


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

def import_rows(valid_rows: List[Dict], resource: str, db: Session) -> Dict:
    """Bulk import rows in a single transaction with per-row savepoints.

    Side effects:
        - materials: creates ``MaterialMovement`` IN if ``current_stock > 0``.
        - products: sets ``version=1`` and ``updated_at``.

    Returns::

        {
            "imported": N,
            "failed": N,
            "errors": [{"row": idx, "field": "...", "message": "..."}],
            "results": [{"row": idx, "status": "imported"|"error", "id": N}],
        }
    """
    imported = 0
    errors = []
    results = []

    for entry in valid_rows:
        data = entry["data"].copy()
        row_idx = entry["row"]

        try:
            with db.begin_nested():
                if resource == "materials":
                    stock = data.pop("current_stock", 0)
                    obj = models.Material(**data)
                    db.add(obj)
                    db.flush()
                    if stock > 0:
                        movement = models.MaterialMovement(
                            material_id=obj.id,
                            quantity=stock,
                            type="IN",
                        )
                        db.add(movement)
                    obj_id = obj.id
                elif resource == "products":
                    data.setdefault("version", 1)
                    data["updated_at"] = datetime.utcnow()
                    obj = models.Product(**data)
                    db.add(obj)
                    db.flush()
                    obj_id = obj.id
                else:  # clients
                    obj = models.Client(**data)
                    db.add(obj)
                    db.flush()
                    obj_id = obj.id

            imported += 1
            results.append({"row": row_idx, "status": "imported", "id": obj_id})

        except Exception as e:
            err_msg = str(e)
            if "UNIQUE constraint" in err_msg and "sku" in err_msg.lower():
                err_msg = f"El SKU '{data.get('sku', '')}' ya existe en la base de datos"
            errors.append({"row": row_idx, "field": "database", "message": err_msg})
            results.append({"row": row_idx, "status": "error", "error": err_msg})

    if imported > 0:
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            return {
                "imported": 0,
                "failed": len(valid_rows),
                "errors": [{"row": -1, "field": "database", "message": f"Error al confirmar la transaccion: {str(e)}"}],
                "results": [r for r in results if r["status"] != "imported"],
            }

    return {
        "imported": imported,
        "failed": len(errors),
        "errors": errors,
        "results": results,
    }


def generate_template(resource: str) -> bytes:
    """Generate an .xlsx template with headers for the given resource.

    Returns raw bytes of the workbook.
    """
    headers = TEMPLATE_HEADERS[resource]
    wb = Workbook()
    ws = wb.active
    ws.title = resource
    ws.append(headers)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
