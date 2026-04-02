from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from database import get_db
import models
from io import BytesIO
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter(prefix="/export", tags=["Export"])


def get_db():
    db = next(get_db())
    try:
        yield db
    finally:
        db.close()


@router.get("/clients")
def export_clients(db: Session = Depends(get_db)):
    """Export all clients to Excel"""
    clients = db.query(models.Client).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Headers
    headers = ["ID", "Nombre", "Email", "Teléfono", "Dirección", "CUIT", "Condición IVA"]
    header_fill = PatternFill(start_color="22382c", end_color="22382c", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, client in enumerate(clients, 2):
        ws.cell(row=row, column=1, value=client.id)
        ws.cell(row=row, column=2, value=client.name or "")
        ws.cell(row=row, column=3, value=client.email or "")
        ws.cell(row=row, column=4, value=client.phone or "")
        ws.cell(row=row, column=5, value=client.address or "")
        ws.cell(row=row, column=6, value=client.cuit or "")
        ws.cell(row=row, column=7, value=client.iva_condition or "")
    
    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
    
    # Save to stream
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clientes.xlsx"}
    )


@router.get("/products")
def export_products(db: Session = Depends(get_db)):
    """Export all products to Excel"""
    products = db.query(models.Product).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Headers
    headers = ["ID", "Nombre", "Descripción", "Precio", "Costo", "Categoría"]
    header_fill = PatternFill(start_color="22382c", end_color="22382c", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.id)
        ws.cell(row=row, column=2, value=product.name or "")
        ws.cell(row=row, column=3, value=product.description or "")
        ws.cell(row=row, column=4, value=product.price or 0)
        ws.cell(row=row, column=5, value=product.cost or 0)
        ws.cell(row=row, column=6, value=product.category or "")
    
    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
    
    # Save to stream
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=productos.xlsx"}
    )


@router.get("/materials")
def export_materials(db: Session = Depends(get_db)):
    """Export all materials to Excel"""
    from sqlalchemy import func
    
    materials = db.query(models.Material).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiales"
    
    # Headers
    headers = ["ID", "SKU", "Nombre", "Categoría", "Costo Unitario", "Stock", "Valor Total"]
    header_fill = PatternFill(start_color="22382c", end_color="22382c", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data with stock calculation
    for row, material in enumerate(materials, 2):
        # Calculate stock
        totals = db.query(
            func.sum(
                func.case(
                    (models.MaterialMovement.type == "IN", models.MaterialMovement.quantity),
                    else_=-models.MaterialMovement.quantity
                )
            )
        ).filter(
            models.MaterialMovement.material_id == material.id
        ).scalar()
        
        stock = totals or 0
        total_value = stock * (material.unit_cost or 0)
        
        ws.cell(row=row, column=1, value=material.id)
        ws.cell(row=row, column=2, value=material.sku or "")
        ws.cell(row=row, column=3, value=material.name or "")
        ws.cell(row=row, column=4, value=material.category or "")
        ws.cell(row=row, column=5, value=material.unit_cost or 0)
        ws.cell(row=row, column=6, value=stock)
        ws.cell(row=row, column=7, value=total_value)
    
    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
    
    # Save to stream
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=materiales.xlsx"}
    )